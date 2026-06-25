# scraper_vendor.py
# Phase 4 — Vendor scraper (a SEPARATE mode from the contractor scraper).
# Finds drywall-material DISTRIBUTORS (the sell-to targets), anchored on the
# prioritized city list with a 20-mi vendor radius. Each discovered place is
# rolled up to its canonical network (GMS/L&W/FBM…) via the vendor alias map and
# tagged record_type='vendor'. Big-box stores are kept but flagged.

import os
from typing import Any, Dict, List, Optional

from agent.schema import GoogleSeed, ContractorRow
from agent.scraper_google import scrape_metro
from agent.vendor import resolve_vendor_network
from agent.lumber import apply_lumber_flag

# Google Maps search phrases for DISTRIBUTORS (sell-to targets), not contractors.
# Distributor-focused so the results lean to suppliers; the vendor-relevance filter
# (vendor.is_distributor) then drops any contractors that still slip through.
VENDOR_QUERIES = [
    "drywall supply",
    "drywall supply company",
    "gypsum supply",
    "drywall distributor",
    "building materials supplier",
    "drywall and acoustical supply",
    "wallboard supply",
]


def discover_vendors(city: str, state: str = "TN", max_charge_usd: Optional[float] = None) -> List[GoogleSeed]:
    """Discover distributors near a city via the Apify Google Maps actor (reuses
    the contractor discovery call with vendor queries + the city's state)."""
    return scrape_metro(city, [], queries=VENDOR_QUERIES, max_charge_usd=max_charge_usd, state=state)


def _as_dict(seed: Any) -> Dict[str, Any]:
    if isinstance(seed, GoogleSeed):
        return seed.model_dump()
    if hasattr(seed, "model_dump"):
        return seed.model_dump()
    return dict(seed)


def build_vendor_row(seed: Any, city_tier: Optional[int] = None,
                     state: Optional[str] = None, client_id: Optional[str] = None,
                     source: str = "google_business") -> Dict[str, Any]:
    """Turn a discovered place (GoogleSeed/dict) into a VENDOR record:
    resolve its canonical network, set vendor_type / is_big_box, and tag it. The
    canonical_network drives roll-up — GMS branches share one network but stay
    distinct per location (canonical_entity_id = network + location, set on save)."""
    s = _as_dict(seed)
    name = s.get("business_name") or ""
    match = resolve_vendor_network(name)
    if match:
        network = match["canonical_network"]
        vendor_type = match["vendor_type"]
    else:
        network = name              # independent → its own network label (groupable)
        vendor_type = "independent"

    row = {
        "business_name": name,
        "city": s.get("city"),
        "zip_code": s.get("zip_code"),
        "address": s.get("address"),
        "phone": s.get("phone"),
        "email": s.get("email"),
        "website": s.get("website"),
        "description": s.get("description"),
        "google_categories": s.get("google_categories") or [],
        "services_listed": s.get("services_listed") or [],
        "google_rating": s.get("google_rating"),
        "google_review_count": s.get("google_review_count"),
        "place_ids": [s["place_id"]] if s.get("place_id") else [],
        # vendor tags
        "record_type": "vendor",
        "vendor_type": vendor_type,
        "is_big_box": vendor_type == "big_box_retailer",
        "canonical_network": network,
        "city_tier": str(city_tier) if city_tier is not None else None,
        "state": state,
        "client_id": client_id,
        "sources": [source],
    }
    # Lumber exclusion (Workstream D) — flag, don't delete.
    return apply_lumber_flag(row)


# ──────────────────────────────────────────────────────────────
# Seed/validation set — Nashville_Drywall_Distributors.xlsx
# ──────────────────────────────────────────────────────────────
def load_seed_distributors(path: Optional[str] = None) -> List[Dict[str, Any]]:
    """Read the provided seed distributor sheet (validation set — confirm + enrich,
    NOT final output). Returns [] gracefully if the file is absent or openpyxl is
    missing. Maps common columns → {business_name, address, phone, city, notes}."""
    path = path or os.getenv("VENDOR_SEED_XLSX", "")
    if not path or not os.path.exists(path):
        print(f"⏩ [vendor-seed] no seed sheet at {path!r} — skipping (provide VENDOR_SEED_XLSX)")
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("⏩ [vendor-seed] openpyxl not installed — skipping seed sheet")
        return []

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip().lower() for h in rows[0]]

    def col(row, *names):
        for n in names:
            if n in headers:
                v = row[headers.index(n)]
                if v not in (None, ""):
                    return str(v).strip()
        return None

    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        name = col(r, "name", "business name", "distributor", "distributor name", "company")
        if not name:
            continue
        out.append({
            "business_name": name,
            "address": col(r, "address", "branch location", "location", "street"),
            "phone": col(r, "phone", "phone number", "telephone"),
            "city": col(r, "city", "town"),
            "notes": col(r, "notes", "note", "comments"),
        })
    print(f"📥 [vendor-seed] loaded {len(out)} seed distributors from {path}")
    return out


def seed_distributor_seeds(path: Optional[str] = None) -> List[GoogleSeed]:
    """Seed distributors (from the validation xlsx) as GoogleSeed objects, so they
    flow through the SAME dedupe → vendor-row → enrich path as discovered vendors.
    place_id is prefixed 'seed:' so the pipeline can tag their source 'vendor_seed'.
    Returns [] when no seed file is configured/present (graceful)."""
    out: List[GoogleSeed] = []
    for i, s in enumerate(load_seed_distributors(path)):
        name = s.get("business_name")
        if not name:
            continue
        out.append(GoogleSeed(
            place_id=f"seed:{i}",
            business_name=name,
            city=s.get("city") or "",
            address=s.get("address"),
            phone=s.get("phone"),
            description=s.get("notes") or "",
        ))
    return out
