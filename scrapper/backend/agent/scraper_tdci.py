# scraper_tdci.py
# Tennessee TDCI statewide license roster — the 2nd TN license source (after the
# Nashville municipal open data), per spec Workstream B priority order.
#
# TDCI (Dept. of Commerce & Insurance, Board for Licensing Contractors) does NOT
# expose a public bulk API — the spec says to obtain a "baseline licensee roster
# export" via an OPEN-RECORDS REQUEST. That request yields a file (CSV/XLSX). So
# this module is a flexible LOADER for that roster file, used as a FALLBACK for
# TN business names the Nashville dataset didn't match (Nashville = Davidson only;
# the TDCI roster is statewide).
#
# Configure the roster file once obtained:
#   TDCI_ROSTER_FILE = /path/to/roster.csv  (or .xlsx, or an http(s) URL)
# Until it's set/present, this returns [] gracefully (license is enrichment-only).
#
# Same contract as dbpr_loader / scraper_tn_license: query_by_normalized_names()
# and fetch_tdci_licenses_for_seeds() → List[DBPRLicense] for the shared matcher.

import csv
import io
import os
from functools import lru_cache
from typing import Any, Dict, List, Optional

import requests

from agent.schema import DBPRLicense, GoogleSeed
from utils.name_normalizer import normalize_name

TDCI_ROSTER_FILE = os.getenv("TDCI_ROSTER_FILE", "").strip()
TIMEOUT = 120

# Header-name hints (case-insensitive substring) → our canonical field. The roster
# column names depend on what TDCI provides, so we auto-detect the common ones.
_COL_HINTS = {
    "name": ("business name", "company", "licensee", "organization", "dba", "name"),
    "classification": ("classification", "license type", "licensetype", "type", "trade", "category"),
    "license_number": ("license number", "license #", "license no", "licenseno", "number", "license_no"),
    "status": ("status",),
    "city": ("city", "town"),
    "state": ("state",),
    "zip_code": ("zip", "postal"),
}
# Optional classification allowlist (comma-separated, case-insensitive). Empty =
# keep all name matches (the Board for Licensing Contractors roster is already
# construction-scoped). Set TDCI_CLASSIFICATIONS to filter by-classification.
_CLASS_ALLOW = tuple(c.strip().lower() for c in os.getenv("TDCI_CLASSIFICATIONS", "").split(",") if c.strip())

_last_scanned = 0


def _pick_columns(headers: List[str]) -> Dict[str, Optional[int]]:
    low = [(h or "").strip().lower() for h in headers]
    out: Dict[str, Optional[int]] = {}
    for field, hints in _COL_HINTS.items():
        idx = None
        for hint in hints:
            for i, h in enumerate(low):
                if hint in h:
                    idx = i
                    break
            if idx is not None:
                break
        out[field] = idx
    return out


def _rows_from_csv(text: str) -> List[List[str]]:
    return list(csv.reader(io.StringIO(text)))


def _rows_from_xlsx(content: bytes) -> List[List[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    return [[("" if c is None else str(c)) for c in row] for row in ws.iter_rows(values_only=True)]


def _load_raw(path: str) -> List[List[str]]:
    """Read the roster file (local path or URL), CSV or XLSX → list of rows."""
    is_xlsx = path.lower().endswith(".xlsx")
    if path.lower().startswith(("http://", "https://")):
        resp = requests.get(path, timeout=TIMEOUT)
        resp.raise_for_status()
        return _rows_from_xlsx(resp.content) if is_xlsx else _rows_from_csv(resp.text)
    if is_xlsx:
        with open(path, "rb") as f:
            return _rows_from_xlsx(f.read())
    with open(path, encoding="utf-8", errors="replace") as f:
        return _rows_from_csv(f.read())


@lru_cache(maxsize=1)
def load_tdci_roster() -> tuple:
    """Parse the configured TDCI roster into normalized license records. Cached per
    process. Returns () gracefully if no file is configured or it can't be read."""
    global _last_scanned
    if not TDCI_ROSTER_FILE:
        return tuple()
    if not TDCI_ROSTER_FILE.lower().startswith(("http://", "https://")) and not os.path.exists(TDCI_ROSTER_FILE):
        print(f"⏩ [TDCI] roster file not found at {TDCI_ROSTER_FILE!r} — skipping (set TDCI_ROSTER_FILE)")
        return tuple()
    try:
        rows = _load_raw(TDCI_ROSTER_FILE)
    except (requests.RequestException, OSError, ValueError, KeyError) as e:
        print(f"⚠️  [TDCI] roster load failed ({e}) — skipping")
        return tuple()
    if not rows:
        return tuple()

    cols = _pick_columns(rows[0])
    if cols.get("name") is None:
        print("⚠️  [TDCI] could not find a name column in roster header — skipping")
        return tuple()

    def cell(row, field):
        i = cols.get(field)
        return (row[i].strip() if i is not None and i < len(row) and row[i] else None)

    out: List[Dict[str, Any]] = []
    for row in rows[1:]:
        name = cell(row, "name")
        if not name:
            continue
        out.append({
            "licensee_name": name,
            "normalized_name": normalize_name(name),
            "license_category": cell(row, "classification") or "",
            "license_status": cell(row, "status") or "registered",
            "license_number": cell(row, "license_number"),
            "city": cell(row, "city"),
            "state": cell(row, "state") or "TN",
            "zip_code": cell(row, "zip_code"),
        })
    _last_scanned = len(out)
    print(f"📁 [TDCI] loaded {len(out)} statewide roster records from {TDCI_ROSTER_FILE}")
    return tuple(out)


def _relevant(category: str) -> bool:
    if not _CLASS_ALLOW:
        return True  # no allowlist → keep all (roster is already contractor-scoped)
    return (category or "").strip().lower() in _CLASS_ALLOW


def _record_to_license(rec: Dict[str, Any]) -> DBPRLicense:
    """Adapt a TDCI roster record to the shared DBPRLicense shape (matcher reuse)."""
    return DBPRLicense(
        license_number=rec.get("license_number") or "",
        license_category=rec.get("license_category") or "",
        licensee_name=rec.get("licensee_name") or "",
        dba_name=None,
        status="Current, Active",
        city=rec.get("city"),
        zip_code=rec.get("zip_code"),
        phone=None,
        original_issue_date=None,
        raw=rec,
    )


def query_by_normalized_names(normalized_names: List[str]) -> List[Dict[str, Any]]:
    """TDCI roster records whose normalized name matches any input (+ optional
    classification allowlist). [] if no roster configured."""
    targets = {n for n in normalized_names if n}
    if not targets:
        return []
    rows = load_tdci_roster()
    if not rows:
        return []
    hits = [r for r in rows if r["normalized_name"] in targets and _relevant(r["license_category"])]
    print(f"✅ [TDCI] scanned {len(rows)} roster rows, matched {len(hits)} ({len(targets)} names)")
    return hits


def fetch_tdci_licenses_for_seeds(seeds: List[GoogleSeed], already=None) -> List[DBPRLicense]:
    """Fallback after Nashville: resolve TN licenses from the TDCI statewide roster
    for the seed names NOT already matched. Returns List[DBPRLicense] (or [] if no
    roster file is configured)."""
    if not load_tdci_roster():
        return []
    names = sorted({s.business_name.strip() for s in seeds if getattr(s, "business_name", None)})
    if not names:
        return []
    matched = {normalize_name(l.licensee_name) for l in (already or [])}
    targets = [normalize_name(n) for n in names if normalize_name(n) not in matched]
    recs = query_by_normalized_names(targets)
    return [_record_to_license(r) for r in recs]


def tdci_count() -> int:
    return _last_scanned
